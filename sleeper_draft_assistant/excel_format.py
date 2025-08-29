from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter

TIER_FILLS = {
    1: "B3E5FC",  # light blue
    2: "C8E6C9",  # light green
    3: "FFE0B2",  # light orange
    4: "F8BBD0",  # light pink
    5: "E1BEE7",  # light purple
}

def apply_tier_colors(ws, tier_col_idx: int, start_row: int = 2):
    # Color rows based on 'tier' column value
    for r in range(start_row, ws.max_row+1):
        cell = ws.cell(row=r, column=tier_col_idx)
        try:
            t = int(cell.value) if cell.value is not None else None
        except:
            t = None
        if t in TIER_FILLS:
            fill = PatternFill(start_color=TIER_FILLS[t], end_color=TIER_FILLS[t], fill_type="solid")
            for c in range(1, ws.max_column+1):
                ws.cell(row=r, column=c).fill = fill

def bold_headers(ws):
    for c in range(1, ws.max_column+1):
        ws.cell(row=1, column=c).font = Font(bold=True)

def autosize(ws):
    widths = {}
    for r in ws.iter_rows(values_only=True):
        for idx, v in enumerate(r, start=1):
            w = len(str(v)) if v is not None else 0
            widths[idx] = max(widths.get(idx, 10), min(40, w+2))
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width
